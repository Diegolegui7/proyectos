import os
import time
import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook

class EcuadorTresPasos(unittest.TestCase):
    def setUp(self):
        self.url = "https://www.chevrolet.com.ec/vans/n400-pasajeros"

        chrome_options = Options()
        # chrome_options.add_argument("--headless")  # descomentar si querés modo sin ventana

        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)

        self.driver.get(self.url)
        self.driver.delete_all_cookies()
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)
        self.wait = WebDriverWait(self.driver, 20)

        # Ruta del Excel
        self.excel_path = r"C:\Users\MH\OneDrive\Documentos\excelresultado.xlsx"

        if os.path.exists(self.excel_path):
            self.wb = load_workbook(self.excel_path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(["Nombre", "Email", "telephone", "Resultado imagen"])

    def test_realizar_acciones(self):
        usuarios = [
            {"nombre": "Diego", "email": "diego@test.com","telephone": "0942345678"},
            {"nombre": "Laura", "email": "laura@test.com","telephone": "0942345678"},
            {"nombre": "Carlos", "email": "carlos@test.com","telephone": "0942345678"}
        ]

        for usuario in usuarios:
            self.paso = 1
            self.driver.get(self.url)
            self.primer_paso()

            try:
                self.wait.until(EC.visibility_of_element_located(
                    (By.XPATH, "//img[contains(@src, '685d3e3151432.jpg')]")))
                resultado_boton = "👀 La imagen está visible"
            except TimeoutException:
                resultado_boton = "❌ La imagen NO está visible (Timeout)"
            except Exception as e:
                resultado_boton = f"Error al encontrar la imagen: {e}"

            print(resultado_boton)

            self.segundo_paso()
            self.tercer_paso(usuario["nombre"], usuario["email"], usuario["telephone"])

            self.ws.append([usuario["nombre"], usuario["email"], resultado_boton])
            self.wb.save(self.excel_path)

            self.driver.switch_to.default_content()
            time.sleep(2)

    def primer_paso(self):
        boton_cotiza = self.wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, ".show-cta-for-specific-breakpoint-up.q-cta-list-item")))
        boton_cotiza.click()

        time.sleep(2)
        iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
        print("Iframes encontrados:", len(iframes))
        for i, f in enumerate(iframes):
            print(f"{i}: {f.get_attribute('src')}")

        self.wait.until(EC.frame_to_be_available_and_switch_to_it(
            (By.XPATH, "//iframe[contains(@src, 'gm_forms/modeloverviews?model=n400')]")
        ))

        self.wait.until(EC.visibility_of_element_located((By.ID, "formulario")))
        self.click_siguiente()

    def segundo_paso(self):
        ciudad = self.wait.until(EC.element_to_be_clickable((By.ID, "city")))
        Select(ciudad).select_by_index(2)

        dealer = self.wait.until(EC.element_to_be_clickable((By.ID, "dealer")))
        Select(dealer).select_by_index(1)

        self.click_siguiente()

    def tercer_paso(self, nombre, email, telephone):
        self.wait.until(EC.element_to_be_clickable((By.ID, "firstname"))).send_keys(nombre)
        self.wait.until(EC.element_to_be_clickable((By.ID, "lastname"))).send_keys("PérezTest")
        self.wait.until(EC.element_to_be_clickable((By.ID, "ci"))).send_keys("1940139031")
        self.wait.until(EC.element_to_be_clickable((By.ID, "telephone"))).send_keys(telephone)
        self.wait.until(EC.element_to_be_clickable((By.ID, "email"))).send_keys(email)

        fecha_estimada = self.wait.until(EC.element_to_be_clickable((By.ID, "estimated-date-purchase")))
        Select(fecha_estimada).select_by_value("1 mes")

        self.wait.until(EC.element_to_be_clickable((By.ID, "icon"))).click()
        boton_enviar = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Enviar']")))
        boton_enviar.click()
        self.wait.until(EC.invisibility_of_element(boton_enviar))
        time.sleep(1)

    def click_siguiente(self):
        if self.paso == 1:
            boton = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Siguiente']")))
        else:
            boton = self.wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[@data-dtm='next:elige tu concesionario']")))
        boton.click()
        self.paso += 1

    def tearDown(self):
        self.driver.quit()
        self.wb.save(self.excel_path)

if __name__ == "__main__":
    unittest.main()


